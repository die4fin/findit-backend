"""
MPL Liquipedia Scraper + Excel Builder — V17
Scrape dari Liquipedia lalu langsung output ke Excel (.xlsx)
dengan kolom Blue_Win yang terdeteksi otomatis dari data-label-type.

Jalankan:
    uv run --with requests --with beautifulsoup4 --with openpyxl scraper_mpl_id_all.py <url> <output_file>

Contoh:
    uv run --with requests --with beautifulsoup4 --with openpyxl scraper_mpl_id_all.py \\
        https://liquipedia.net/mobilelegends/MPL/Indonesia/Season_17/Regular_Season \\
        mpl_id_s17.xlsx
"""

import re
import argparse
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────── CLI ARGS ────────────────────────────
parser = argparse.ArgumentParser(
    description="Scrape MPL data dari Liquipedia dan simpan ke Excel."
)
parser.add_argument(
    "url",
    help="URL halaman Liquipedia yang ingin di-scrape. "
         "Contoh: https://liquipedia.net/mobilelegends/MPL/Indonesia/Season_17/Regular_Season"
)
parser.add_argument(
    "output_file",
    help="Nama file Excel output, harus berekstensi .xlsx. Contoh: mpl_id_s17.xlsx"
)
args = parser.parse_args()

URL         = args.url
OUTPUT_FILE = args.output_file
HEADERS     = {"User-Agent": "MPL_Master_Scraper/17.0 (belajar_bareng@gmail.com)"}


# ─────────────────────────── HELPERS ─────────────────────────────
def get_match_score(match_div):
    """Ambil skor akhir match (misal 2:1) dari dua div.brkts-matchlist-score."""
    cells = [
        s for s in match_div.find_all(
            'div', class_=lambda x: x and 'brkts-matchlist-score' in x
        )
        if s.parent == match_div
    ]
    if len(cells) >= 2:
        s1 = cells[0].get_text(strip=True)
        s2 = cells[1].get_text(strip=True)
        if s1.isdigit() and s2.isdigit():
            return int(s1), int(s2)
    return None, None


def get_team_names(match_div):
    """Nama panjang tim dari matchlist (misal 'Bigetron by Vitality')."""
    opponents = match_div.find_all(
        'div', class_=lambda x: x and 'brkts-matchlist-opponent' in x
    )
    t1, t2 = "Tim Kiri", "Tim Kanan"
    if len(opponents) >= 2:
        a1 = opponents[0].find('a', title=True)
        t1 = a1['title'].split(' (')[0] if a1 else opponents[0].get_text(strip=True)[:20]
        a2 = opponents[-1].find('a', title=True)
        t2 = a2['title'].split(' (')[0] if a2 else opponents[-1].get_text(strip=True)[:20]
    return t1, t2


def get_team_short(match_div):
    """
    Short name (misalnya 'BTR', 'AE') dari span.name di matchlist.
    Return: (short_kiri, short_kanan)
    """
    opponents = match_div.find_all(
        'div', class_=lambda x: x and 'brkts-matchlist-opponent' in x
    )
    def short_from(opp):
        # span.visible-xs biasanya berisi nama pendek
        span = opp.find('span', class_=lambda x: x and 'visible-xs' in x)
        if span:
            t = span.get_text(strip=True)
            if t:
                return t
        # Fallback: ambil teks dari span.name
        span2 = opp.find('span', class_='name')
        if span2:
            return span2.get_text(strip=True)
        a = opp.find('a', title=True)
        return a['title'].split(' (')[0] if a else "?"

    if len(opponents) >= 2:
        return short_from(opponents[0]), short_from(opponents[-1])
    return "Tim Kiri", "Tim Kanan"


def get_hero_color(a_tag):
    """Warna sisi hero dari class parent: --blue atau --red."""
    classes = ' '.join(a_tag.parent.get('class', []))
    if 'brkts-popup-side-color--blue' in classes:
        return 'blue'
    if 'brkts-popup-side-color--red' in classes:
        return 'red'
    return 'default'


def parse_games(popup):
    """
    Parse setiap game dari popup. Return list of dict per game:
      kiri_picks, kanan_picks, warna_kiri, warna_kanan,
      kiri_wins (bool | None)   ← dari data-label-type

    Deduplication: Liquipedia render setiap game 2x (desktop+mobile).
    """
    games = []
    seen = set()

    rows = popup.find_all(
        'div', class_=lambda x: x and 'brkts-popup-body-grid-row' in x
    )
    for row in rows:
        # ── Picks ──
        kiri_div = row.find(
            'div',
            class_=lambda x: x
                and 'brkts-champion-icon' in x
                and 'brkts-popup-body-element-thumbs-right' not in x
        )
        kanan_div = row.find(
            'div',
            class_=lambda x: x and 'brkts-popup-body-element-thumbs-right' in x
        )
        if not kiri_div or not kanan_div:
            continue

        kiri_links  = kiri_div.find_all('a', title=True)
        kanan_links = kanan_div.find_all('a', title=True)
        if len(kiri_links) < 5 or len(kanan_links) < 5:
            continue

        kiri_titles  = tuple(a['title'] for a in kiri_links[:5])
        kanan_titles = tuple(a['title'] for a in kanan_links[:5])
        key = (kiri_titles, kanan_titles)
        if key in seen:
            continue
        seen.add(key)

        # ── Warna ──
        warna_kiri  = get_hero_color(kiri_links[0])
        warna_kanan = get_hero_color(kanan_links[0])

        # ── Pemenang per game ──
        # generic-label PERTAMA di row = hasil tim kiri (result-win / result-loss)
        label = row.find('div', class_=lambda x: x and 'generic-label' in x)
        label_type = label.get('data-label-type', '') if label else ''

        if label_type == 'result-win':
            kiri_wins = True
        elif label_type == 'result-loss':
            kiri_wins = False
        else:
            kiri_wins = None  # belum dimainkan / tidak ada data

        games.append({
            'kiri_picks':  list(kiri_titles),
            'kanan_picks': list(kanan_titles),
            'warna_kiri':  warna_kiri,
            'warna_kanan': warna_kanan,
            'kiri_wins':   kiri_wins,
        })
    return games


def parse_bans(popup):
    """
    Parse ban per game dari tabel brkts-popup-mapveto.
    Return list of {'kiri': [...], 'kanan': [...]}
    """
    result = []
    mapveto = popup.find('div', class_=lambda x: x and 'brkts-popup-mapveto' in x)
    if not mapveto:
        return result

    ban_rows = mapveto.find_all(
        'tr', class_=lambda x: x and 'brkts-popup-mapveto__ban-round' in x
    )
    for row in ban_rows:
        tds = row.find_all(
            'td', class_=lambda x: x and 'brkts-popup-mapveto__ban-round-picks' in x
        )
        if len(tds) < 2:
            continue
        result.append({
            'kiri':  [a['title'] for a in tds[0].find_all('a', title=True)],
            'kanan': [a['title'] for a in tds[1].find_all('a', title=True)],
        })
    return result


def pad(lst, n=5):
    return list(lst)[:n] + [''] * (n - len(lst))


# ─────────────────────────── SCRAPE ──────────────────────────────
print("Scraping Dataset...\n")

response = requests.get(URL, headers=HEADERS)
response.raise_for_status()
soup = BeautifulSoup(response.text, 'html.parser')

all_matches = soup.find_all('div', class_=lambda x: x and 'brkts-matchlist-match' in x)
print(f"Total slot pertandingan ditemukan: {len(all_matches)}")

data_rows = []   # list of dict untuk tiap game

for match in all_matches:
    popup = match.find('div', class_=lambda x: x and 'brkts-popup' in x)
    if not popup:
        continue

    # Skip pertandingan yang belum dimainkan (tidak ada durasi)
    if not popup.find_all(string=re.compile(r'^\s*\d{2}:\d{2}\s*$')):
        continue

    team_long1, team_long2 = get_team_names(match)
    short_kiri, short_kanan = get_team_short(match)
    match_title = f"{team_long1} vs {team_long2}"

    games = parse_games(popup)
    bans  = parse_bans(popup)

    for g_idx, game in enumerate(games):
        ban = bans[g_idx] if g_idx < len(bans) else {'kiri': [], 'kanan': []}

        # Tentukan sisi blue/red
        if game['warna_kiri'] == 'blue':
            blue_side, red_side       = 'kiri', 'kanan'
            blue_team, red_team       = short_kiri, short_kanan
        else:
            blue_side, red_side       = 'kanan', 'kiri'
            blue_team, red_team       = short_kanan, short_kiri

        # Blue_Win: 1 jika blue team menang game ini, 0 jika kalah
        kiri_wins = game['kiri_wins']
        if kiri_wins is None:
            blue_win = ''
        elif blue_side == 'kiri':
            blue_win = 1 if kiri_wins else 0
        else:
            blue_win = 0 if kiri_wins else 1

        b_picks = pad(game[f'{blue_side}_picks'])
        r_picks = pad(game[f'{red_side}_picks'])
        b_bans  = pad(ban[blue_side])
        r_bans  = pad(ban[red_side])

        row = {
            'Match_Title': match_title,
            'Game_Number': g_idx + 1,
            'Blue_Team':   blue_team,
            'Red_Team':    red_team,
            'B_B1': b_bans[0], 'B_B2': b_bans[1], 'B_B3': b_bans[2],
            'B_B4': b_bans[3], 'B_B5': b_bans[4],
            'R_B1': r_bans[0], 'R_B2': r_bans[1], 'R_B3': r_bans[2],
            'R_B4': r_bans[3], 'R_B5': r_bans[4],
            'B_P1': b_picks[0], 'B_P2': b_picks[1], 'B_P3': b_picks[2],
            'B_P4': b_picks[3], 'B_P5': b_picks[4],
            'R_P1': r_picks[0], 'R_P2': r_picks[1], 'R_P3': r_picks[2],
            'R_P4': r_picks[3], 'R_P5': r_picks[4],
            'Blue_Win': blue_win,
        }
        data_rows.append(row)

print(f"Total game terkumpul: {len(data_rows)}")


# ─────────────────────────── BUILD EXCEL ─────────────────────────
COLS = [
    'Match_Title', 'Game_Number', 'Blue_Team', 'Red_Team',
    'B_B1','B_B2','B_B3','B_B4','B_B5',
    'R_B1','R_B2','R_B3','R_B4','R_B5',
    'B_P1','B_P2','B_P3','B_P4','B_P5',
    'R_P1','R_P2','R_P3','R_P4','R_P5',
    'Blue_Win',
]

COL_WIDTHS = {
    'Match_Title': 24, 'Game_Number': 7,
    'Blue_Team': 12, 'Red_Team': 12,
    **{f'B_B{i}': 13 for i in range(1,6)},
    **{f'R_B{i}': 13 for i in range(1,6)},
    **{f'B_P{i}': 13 for i in range(1,6)},
    **{f'R_P{i}': 13 for i in range(1,6)},
    'Blue_Win': 9,
}

# Palette
C_HDR_BG   = '1F3864'; C_HDR_FG  = 'FFFFFF'
C_META_BG  = 'F2F2F2'
C_BBBAN_BG = 'BDD7EE'          # blue team ban — biru muda
C_RBBAN_BG = 'FADADD'          # red team ban  — merah muda
C_BPICK_BG = '2E75B6'; C_BPICK_FG = 'FFFFFF'   # blue pick
C_RPICK_BG = 'C00000'; C_RPICK_FG = 'FFFFFF'   # red pick
C_WIN_BG   = 'E2EFDA'          # blue_win = 1
C_LOSE_BG  = 'FCE4D6'          # blue_win = 0

thin   = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style(cell, bg=None, fg='000000', bold=False, center=True):
    if bg:
        cell.fill = PatternFill('solid', start_color=bg)
    cell.font      = Font(name='Arial', size=10, bold=bold, color=fg)
    cell.alignment = Alignment(
        horizontal='center' if center else 'left',
        vertical='center', wrap_text=True
    )
    cell.border = border

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "MPL ID S17"
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

# Header
for ci, col in enumerate(COLS, 1):
    c = ws.cell(row=1, column=ci, value=col)
    style(c, bg=C_HDR_BG, fg=C_HDR_FG, bold=True)
ws.row_dimensions[1].height = 28

for ci, col in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 12)

# Data rows
for ri, row in enumerate(data_rows, 2):
    bw = row.get('Blue_Win', '')
    for ci, col in enumerate(COLS, 1):
        val = row.get(col, '')
        if col == 'Game_Number' and val != '':
            val = int(val)
        if col == 'Blue_Win' and val != '':
            val = int(val)

        c = ws.cell(row=ri, column=ci, value=val)

        if col in ('Match_Title', 'Blue_Team', 'Red_Team'):
            style(c, bg=C_META_BG, center=False)
        elif col == 'Game_Number':
            style(c, bg=C_META_BG, bold=True)
        elif col.startswith('B_B'):
            style(c, bg=C_BBBAN_BG)
        elif col.startswith('R_B'):
            style(c, bg=C_RBBAN_BG)
        elif col.startswith('B_P'):
            style(c, bg=C_BPICK_BG, fg=C_BPICK_FG, bold=True)
        elif col.startswith('R_P'):
            style(c, bg=C_RPICK_BG, fg=C_RPICK_FG, bold=True)
        elif col == 'Blue_Win':
            if bw == 1:
                style(c, bg=C_WIN_BG, bold=True)
            elif bw == 0:
                style(c, bg=C_LOSE_BG)
            else:
                style(c)

    ws.row_dimensions[ri].height = 22

wb.save(OUTPUT_FILE)
print(f"\n✅ Excel tersimpan: {OUTPUT_FILE}  ({len(data_rows)} game, {len(data_rows)//len(set(r['Match_Title'] for r in data_rows))} avg game/match)")
